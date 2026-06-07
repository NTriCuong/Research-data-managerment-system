from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from uuid import UUID
import re

from app.core.exceptions import BadRequestException
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.auth.user import User
from app.models.core.core_research_object import CoreResearchObject
from app.models.enum import AccessLevel, AuthorRole, WorkflowStatus
from app.models.reference.department import Department
from app.models.reference.keyword import Keyword
from app.models.reference.output_type import OutputType
from app.models.reference.research_domain import ResearchDomain
from app.models.reference.researcher import Researcher
from app.models.staging.stg_file_attachment import StgFileAttachment
from app.models.staging.stg_research_object import StgResearchObject
from app.models.staging.stg_research_object_author import StgResearchObjectAuthor
from app.models.staging.stg_research_object_domain import StgResearchObjectDomain
from app.models.staging.stg_research_object_keyword import StgResearchObjectKeyword
from app.schemas.imports import ImportRowError, StagingExcelImportResponse
from app.schemas.files import IncomingFile
from app.services.logs.audit_service import audit_service
from app.services.logs.workflow_service import workflow_service


@dataclass(slots=True)
class ParsedWorkbook:
    sheets: dict[str, list[dict[str, str]]]


@dataclass(slots=True)
class ReferenceMaps:
    output_types: dict[str, OutputType]
    departments: dict[str, Department]
    domains: dict[str, ResearchDomain]
    keywords: dict[str, Keyword]
    researchers: dict[str, Researcher]
    existing_titles: set[str]
    existing_identifiers: set[str]


class StagingExcelImportService:
    async def import_workbook(
        self,
        db: AsyncSession,
        *,
        file: IncomingFile,
        current_user: User,
    ) -> StagingExcelImportResponse:
        workbook = await self._read_workbook(file)
        project_rows = workbook.sheets.get("researchrecords", [])
        file_rows = workbook.sheets.get("fileattachments", [])
        if not project_rows:
            raise BadRequestException("Tệp Excel phải chứa sheet projects")

        refs = await self._preload_references(db)
        errors: list[ImportRowError] = []
        imported_projects: dict[str, StgResearchObject] = {}
        imported_file_count = 0
        workbook_titles: set[str] = set()
        workbook_identifiers: set[str] = set()

        for row_index, row in enumerate(project_rows, start=2):
            try:
                import_key = self._required(row, "import_key")
                if import_key in imported_projects:
                    raise ValueError(f"import_key bị trùng: {import_key}")

                self._validate_project_uniqueness(row, refs, workbook_titles, workbook_identifiers)
                obj = self._build_project(row=row, refs=refs, current_user=current_user)
                db.add(obj)
                await db.flush()

                domains, keywords, authors = await self._add_project_relations(db, obj=obj, row=row, refs=refs)
                await workflow_service.write_history(
                    db,
                    staging_id=obj.staging_id,
                    performed_by=current_user.user_id,
                    from_status=None,
                    to_status=WorkflowStatus.draft,
                    action_code="IMPORT_EXCEL_DRAFT",
                    action_note=f"Excel import key: {import_key}",
                )
                self._set_quality_score(obj, domain_count=len(domains), keyword_count=len(keywords), author_count=len(authors))
                imported_projects[import_key] = obj
            except ValueError as exc:
                errors.append(ImportRowError(sheet="projects", row_number=row_index, message=str(exc)))

        for row_index, row in enumerate(file_rows, start=2):
            try:
                import_key = self._required(row, "import_key")
                project = imported_projects.get(import_key)
                if project is None:
                    raise ValueError(f"Không tìm thấy import_key cho tệp đính kèm: {import_key}")
                db.add(self._build_file_attachment(row=row, staging_id=project.staging_id, current_user=current_user))
                imported_file_count += 1
            except ValueError as exc:
                errors.append(ImportRowError(sheet="file_attachments", row_number=row_index, message=str(exc)))

        if errors:
            await db.rollback()
            return StagingExcelImportResponse(
                imported_projects=0,
                imported_file_attachments=0,
                staging_ids=[],
                errors=errors,
            )

        await db.flush()
        await audit_service.write_log(
            db,
            actor_user_id=current_user.user_id,
            action_code="IMPORT_EXCEL_STAGING_DRAFTS",
            target_schema="staging",
            target_table="research_objects",
            target_id=None,
            new_value={"projects": len(imported_projects), "file_attachments": imported_file_count},
            message="Imported staging drafts from Excel",
        )
        return StagingExcelImportResponse(
            imported_projects=len(imported_projects),
            imported_file_attachments=imported_file_count,
            staging_ids=[obj.staging_id for obj in imported_projects.values()],
        )

    async def _preload_references(self, db: AsyncSession) -> ReferenceMaps:
        output_types = (await db.execute(select(OutputType).where(OutputType.is_active.is_(True)))).scalars().all()
        departments = (await db.execute(select(Department).where(Department.is_active.is_(True)))).scalars().all()
        domains = (await db.execute(select(ResearchDomain).where(ResearchDomain.is_active.is_(True)))).scalars().all()
        keywords = (await db.execute(select(Keyword))).scalars().all()
        researchers = (await db.execute(select(Researcher))).scalars().all()
        staging_objects = (
            await db.execute(select(StgResearchObject.title, StgResearchObject.identifier).where(StgResearchObject.deleted_at.is_(None)))
        ).all()
        core_objects = (
            await db.execute(select(CoreResearchObject.title, CoreResearchObject.identifier).where(CoreResearchObject.deleted_at.is_(None)))
        ).all()

        return ReferenceMaps(
            output_types=self._map_output_types(output_types),
            departments=self._map_departments(departments),
            domains=self._map_domains(domains),
            keywords=self._map_keywords(keywords),
            researchers=self._map_researchers(researchers),
            existing_titles={self._norm_key(title) for title, _ in [*staging_objects, *core_objects] if title},
            existing_identifiers={self._norm_key(identifier) for _, identifier in [*staging_objects, *core_objects] if identifier},
        )

    def _build_project(self, *, row: dict[str, str], refs: ReferenceMaps, current_user: User) -> StgResearchObject:
        output_type = self._resolve_from_map(refs.output_types, self._required(row, "output_type"), "Loại sản phẩm")
        department = self._resolve_from_map(refs.departments, self._required(row, "department"), "Đơn vị")
        year = self._optional_int(row, "year")
        if year is not None and not 1900 <= year <= 2100:
            raise ValueError("year phải nằm trong khoảng từ 1900 đến 2100")

        return StgResearchObject(
            title=self._required(row, "title"),
            output_type_id=output_type.output_type_id,
            department_id=department.department_id,
            year=year,
            description=self._optional(row, "description"),
            abstract=self._optional(row, "abstract"),
            start_date=self._optional_date(row, "start_date"),
            end_date=self._optional_date(row, "end_date"),
            date_issued=self._optional_date(row, "date_issued"),
            publisher=self._optional(row, "publisher"),
            language=self._optional(row, "language") or "vi",
            identifier=self._optional(row, "identifier"),
            external_url=self._optional(row, "external_url"),
            source=self._optional(row, "source"),
            relation=self._optional(row, "relation"),
            coverage=self._optional(row, "coverage"),
            rights=self._optional(row, "rights"),
            access_level=AccessLevel.internal,
            workflow_status=WorkflowStatus.draft,
            created_by=current_user.user_id,
        )

    async def _add_project_relations(
        self,
        db: AsyncSession,
        *,
        obj: StgResearchObject,
        row: dict[str, str],
        refs: ReferenceMaps,
    ) -> tuple[list[ResearchDomain], list[Keyword], list[Researcher]]:
        domains = [self._resolve_from_map(refs.domains, value, "Lĩnh vực nghiên cứu") for value in self._split(row.get("domains"))]
        keywords = await self._resolve_or_create_keywords(db, self._split(row.get("keywords")), refs)
        authors = await self._resolve_or_create_researchers(db, self._split(row.get("authors")), refs)

        db.add_all([StgResearchObjectDomain(staging_id=obj.staging_id, domain_id=x.domain_id) for x in domains])
        db.add_all([StgResearchObjectKeyword(staging_id=obj.staging_id, keyword_id=x.keyword_id) for x in keywords])
        db.add_all(
            [
                StgResearchObjectAuthor(
                    staging_id=obj.staging_id,
                    researcher_id=author.researcher_id,
                    full_name=author.full_name,
                    email=author.email,
                    affiliation=None,
                    author_order=index,
                    author_role=AuthorRole.creator if index == 1 else AuthorRole.contributor,
                )
                for index, author in enumerate(authors, start=1)
            ]
        )
        await db.flush()
        return domains, keywords, authors

    async def _resolve_or_create_keywords(self, db: AsyncSession, values: list[str], refs: ReferenceMaps) -> list[Keyword]:
        keywords: list[Keyword] = []
        for value in values:
            key = self._norm_key(value)
            keyword = refs.keywords.get(key)
            if keyword is None:
                keyword = Keyword(keyword_text=value, normalized_text=key)
                db.add(keyword)
                await db.flush()
                refs.keywords[str(keyword.keyword_id)] = keyword
                refs.keywords[key] = keyword
            keywords.append(keyword)
        return keywords

    async def _resolve_or_create_researchers(self, db: AsyncSession, values: list[str], refs: ReferenceMaps) -> list[Researcher]:
        if not values:
            raise ValueError("authors là trường bắt buộc")
        researchers: list[Researcher] = []
        for value in values:
            parsed = self._parse_author(value)
            keys = [self._norm_key(item) for item in (parsed["researcher_id"], parsed["email"], parsed["code"], parsed["full_name"]) if item]
            researcher = next((refs.researchers[key] for key in keys if key in refs.researchers), None)
            if researcher is None:
                researcher = Researcher(
                    researcher_id=self._parse_uuid(parsed["researcher_id"]) if parsed["researcher_id"] else None,
                    full_name=parsed["full_name"] or parsed["email"] or value,
                    email=parsed["email"],
                    researcher_code=parsed["code"],
                    is_internal=True,
                )
                db.add(researcher)
                await db.flush()
                self._index_researcher(refs.researchers, researcher)
            researchers.append(researcher)
        return researchers

    def _build_file_attachment(self, *, row: dict[str, str], staging_id: UUID, current_user: User) -> StgFileAttachment:
        original_filename = self._required(row, "original_filename")
        storage_path = self._required(row, "storage_path")
        return StgFileAttachment(
            staging_id=staging_id,
            original_filename=original_filename,
            stored_filename=self._optional(row, "stored_filename") or Path(original_filename).name,
            storage_path=storage_path,
            mime_type=self._optional(row, "mime_type") or "application/octet-stream",
            file_extension=self._optional(row, "file_extension") or Path(original_filename).suffix[:20] or None,
            file_size_bytes=self._positive_int(row, "file_size_bytes"),
            checksum_sha256=self._optional(row, "checksum_sha256"),
            uploaded_by=current_user.user_id,
            access_level=AccessLevel.internal,
        )

    async def _read_workbook(self, file: IncomingFile) -> ParsedWorkbook:
        content = file.content
        if not content:
            raise BadRequestException("Tệp Excel tải lên đang trống")
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        except (InvalidFileException, OSError, ValueError) as exc:
            raise BadRequestException("Tệp .xlsx không hợp lệ") from exc

        sheets: dict[str, list[dict[str, str]]] = {}
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                sheets[self._normalize_header(sheet_name)] = []
                continue
            headers = [self._normalize_header(self._to_cell_text(value)) for value in rows[0]]
            sheets[self._normalize_header(sheet_name)] = [
                {
                    headers[index]: self._to_cell_text(value)
                    for index, value in enumerate(row)
                    if index < len(headers) and headers[index] and self._to_cell_text(value)
                }
                for row in rows[1:]
                if any(self._to_cell_text(value) for value in row)
            ]
        workbook.close()
        return ParsedWorkbook(sheets=sheets)

    def _validate_project_uniqueness(
        self,
        row: dict[str, str],
        refs: ReferenceMaps,
        workbook_titles: set[str],
        workbook_identifiers: set[str],
    ) -> None:
        title = self._required(row, "title")
        title_key = self._norm_key(title)
        if title_key in refs.existing_titles or title_key in workbook_titles:
            raise ValueError(f"Tiêu đề bị trùng: {title}")
        workbook_titles.add(title_key)

        identifier = self._optional(row, "identifier")
        if identifier:
            identifier_key = self._norm_key(identifier)
            if identifier_key in refs.existing_identifiers or identifier_key in workbook_identifiers:
                raise ValueError(f"Định danh bị trùng: {identifier}")
            workbook_identifiers.add(identifier_key)

    def _set_quality_score(
        self,
        obj: StgResearchObject,
        *,
        domain_count: int,
        keyword_count: int,
        author_count: int,
    ) -> None:
        checks = {
            "title": bool(obj.title and obj.title.strip()),
            "output_type_id": obj.output_type_id is not None,
            "department_id": obj.department_id is not None,
            "year": obj.year is not None,
            "authors": author_count > 0,
            "domains": domain_count > 0,
            "keywords": keyword_count > 0,
        }
        passed = sum(1 for ok in checks.values() if ok)
        score = (Decimal(passed) * Decimal("100")) / Decimal(len(checks))
        obj.metadata_quality_score = score.quantize(Decimal("0.01"))
        obj.metadata_quality_detail = {"passed": passed, "total": len(checks), "checks": checks}

    def _map_output_types(self, rows: list[OutputType]) -> dict[str, OutputType]:
        result: dict[str, OutputType] = {}
        for row in rows:
            self._put(result, str(row.output_type_id), row)
            self._put(result, row.type_code, row)
            self._put(result, row.type_name, row)
        return result

    def _map_departments(self, rows: list[Department]) -> dict[str, Department]:
        result: dict[str, Department] = {}
        for row in rows:
            self._put(result, str(row.department_id), row)
            self._put(result, row.department_code, row)
            self._put(result, row.department_name, row)
        return result

    def _map_domains(self, rows: list[ResearchDomain]) -> dict[str, ResearchDomain]:
        result: dict[str, ResearchDomain] = {}
        for row in rows:
            self._put(result, str(row.domain_id), row)
            self._put(result, row.domain_code, row)
            self._put(result, row.domain_name, row)
        return result

    def _map_keywords(self, rows: list[Keyword]) -> dict[str, Keyword]:
        result: dict[str, Keyword] = {}
        for row in rows:
            self._put(result, str(row.keyword_id), row)
            self._put(result, row.keyword_text, row)
            self._put(result, row.normalized_text, row)
        return result

    def _map_researchers(self, rows: list[Researcher]) -> dict[str, Researcher]:
        result: dict[str, Researcher] = {}
        for row in rows:
            self._index_researcher(result, row)
        return result

    def _index_researcher(self, target: dict[str, Researcher], row: Researcher) -> None:
        self._put(target, str(row.researcher_id), row)
        self._put(target, row.email, row)
        self._put(target, row.researcher_code, row)
        self._put(target, row.full_name, row)

    def _put(self, target: dict[str, object], value: str | None, obj: object) -> None:
        if value:
            target[self._norm_key(value)] = obj

    def _resolve_from_map(self, mapping: dict[str, object], value: str, label: str):
        obj = mapping.get(self._norm_key(value))
        if obj is None:
            raise ValueError(f"Không tìm thấy {label.lower()}: {value}")
        return obj

    def _parse_author(self, value: str) -> dict[str, str | None]:
        uuid_value = self._parse_uuid(value)
        if uuid_value:
            return {"researcher_id": str(uuid_value), "email": None, "code": None, "full_name": None}

        match = re.fullmatch(r"(?P<name>.+?)\s*<(?P<email>[^<>@\s]+@[^<>@\s]+)>", value)
        if match:
            return {"researcher_id": None, "email": match.group("email"), "code": None, "full_name": match.group("name").strip()}
        if self._looks_like_email(value):
            return {"researcher_id": None, "email": value, "code": None, "full_name": value.split("@", 1)[0]}
        return {"researcher_id": None, "email": None, "code": None, "full_name": value}

    def _split(self, value: str | None) -> list[str]:
        if not value:
            return []
        items: list[str] = []
        seen: set[str] = set()
        for item in re.split(r"[;,|]", value):
            cleaned = item.strip()
            normalized = cleaned.lower()
            if cleaned and normalized not in seen:
                items.append(cleaned)
                seen.add(normalized)
        return items

    def _required(self, row: dict[str, str], key: str) -> str:
        value = self._optional(row, key)
        if not value:
            raise ValueError(f"{key} là trường bắt buộc")
        return value

    def _optional(self, row: dict[str, str], key: str) -> str | None:
        value = row.get(key)
        return value.strip() if value and value.strip() else None

    def _optional_int(self, row: dict[str, str], key: str) -> int | None:
        value = self._optional(row, key)
        if value is None:
            return None
        try:
            return int(Decimal(value))
        except InvalidOperation as exc:
            raise ValueError(f"{key} phải là một số") from exc

    def _positive_int(self, row: dict[str, str], key: str) -> int:
        value = self._optional_int(row, key)
        if value is None or value <= 0:
            raise ValueError(f"{key} phải lớn hơn 0")
        return value

    def _optional_date(self, row: dict[str, str], key: str) -> date | None:
        value = self._optional(row, key)
        if value is None:
            return None
        try:
            return date.fromisoformat(value)
        except ValueError as exc:
            raise ValueError(f"{key} phải sử dụng định dạng YYYY-MM-DD") from exc

    def _normalize_header(self, value: str) -> str:
        return re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")

    def _norm_key(self, value: str) -> str:
        return re.sub(r"\s+", " ", value.strip().lower())

    def _to_cell_text(self, value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _looks_like_email(self, value: str) -> bool:
        return bool(re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", value))

    def _parse_uuid(self, value: str | None) -> UUID | None:
        if not value:
            return None
        try:
            return UUID(value)
        except ValueError:
            return None


staging_excel_import_service = StagingExcelImportService()
